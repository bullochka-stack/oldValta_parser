import openpyxl
from selenium import webdriver
from selenium.common.exceptions import ElementNotInteractableException, NoSuchElementException
import time
import os.path

PATH = "./msedgedriver.exe"


def make_excel_file():
    columns = ('НАИМЕНОВАНИЕ', 'АРТИКУЛ', 'БРЕНД', 'КАТЕГОРИЯ', 'ШТРИХ-КОД', 'ОПИСАНИЕ', 'ХАРАКТЕРИСТИКИ', 'ССЫЛКА НА ФОТО')
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_sheet.title = "old.valta.ru parser"
    k = len(columns)
    for i in range(1, k+1):
        my_sheet.cell(row=1, column=i).value = columns[i-1]

    my_wb.save('./parse.xlsx')


def get_info():
    with open('url_list.txt') as file:
        lines = [line.strip() for line in file.readlines()]
        driver = webdriver.Edge(PATH)

        score = 1

        for line in lines:
            score += 1
            driver.get(line)
            time.sleep(1)
            try:
                driver.find_element_by_xpath("//*[@id='form-btn-region']/span").click()
            except ElementNotInteractableException:
                pass
            time.sleep(1)

            artikul = (driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[1]/div[2]/span').text).lstrip("Артикул: ")
            name = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[1]/div[1]/b').text
            brand = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div[1]/a').text
            category = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div[3]/span/a').text
            kod = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div[7]/span').text
            description = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[4]/div[2]').text
            try:
                driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[5]/div[1]/div[2]/div[1]/span[1]').click()
                characteristics = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[5]/div[1]/div[2]/div[2]').text
            except NoSuchElementException:
                characteristics = '-'

            link_to_photo = driver.find_element_by_xpath('//*[@id="content"]/div[1]/div[2]/div[2]/div/div/div[2]/div[1]/div[2]/ul/li/a/img').get_attribute('src')
            input_info_in_file(name, artikul, brand, category, kod, description, characteristics, link_to_photo)
            print(f"{score-1} complete!")


def input_info_in_file(name, artikul, brand, category, kod, description, characteristics, link_to_photo):
    wb = openpyxl.load_workbook(filename='./parse.xlsx')
    sheet = wb['old.valta.ru parser']
    data_of_product = (name, artikul, brand, category, kod, description, characteristics, link_to_photo)
    k =len(data_of_product)
    for i in range(1, k+1):
        sheet.cell(row=score, column=i).value = data_of_product[i-1]
    wb.save('./parse.xlsx')




if os.path.exists('./parse.xlsx') == False:
    make_excel_file()
    get_info()
else:
    print("File exist")
    get_info()
